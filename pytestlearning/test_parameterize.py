import pytest
import allure
from allure_commons.types import AttachmentType
import openpyxl


def get_data():
    return [

        ["trainer@way2automation.com", "kjsdfbksdf"],
        ["java@way2automation.com", "sdf"],
        ["info@way2automation.com", "sdfsdf"]

    ]
    # workbook = openpyxl.load_workbook("..//excel//testdata.xlsx")
    # sheet = workbook["LoginTest"]
    # totalrows = sheet.max_row
    # totalcols = sheet.max_column
    # mainList = []
    #
    # for i in range(2, totalrows+1):
    #     dataList = []
    #     for j in range(1, totalcols+1):
    #        data = sheet.cell(row=i,column=j).value
    #        dataList.insert(j,data)
    #     mainList.insert(i,dataList)
    # return mainList
    #
    #
    #


@allure.feature("Login Test")
@allure.severity(allure.severity_level.MINOR)
@pytest.mark.usefixtures("log_on_failure")
@pytest.mark.parametrize("username,password", get_data())
def test_dologin(page, username, password):

    with allure.step(f"Enter username and password {username},and {password}"):
        page.fill("#email",username)
        page.fill("#pass",password)
        print(username, "---", password)
        #assert 1 == 2

    #allure.attach(page.screenshot(path="screenshot/fullpage.png"),name="dologin",attachment_type=AttachmentType.PNG)
